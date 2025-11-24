#!/usr/bin/env python3
"""
Cari kartlarını Excel dosyasından SQLite veritabanına aktar
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))

import openpyxl
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from aliaport_api.modules.cari.models import Cari

# Veritabanı bağlantısı
DATABASE_URL = "sqlite:///./aliaport_api/database.db"
engine = create_engine(DATABASE_URL)

# Tabloları oluştur
from aliaport_api.config.database import Base
Base.metadata.create_all(bind=engine)

Session = sessionmaker(bind=engine)
session = Session()

def import_cari_kartlari():
    """Excel dosyasından cari kartlarını oku ve veritabanına ekle"""
    
    excel_file = 'c:\\Aliaport\\Aliaport_v3_1\\kartlarexcel\\cari_sablon_PARSED_SAFE.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"❌ Dosya bulunamadı: {excel_file}")
        return
    
    # Excel dosyasını aç
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Başlıkları oku (1. satır)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"📋 Başlıklar: {headers}")
    print(f"📊 Toplam satır: {ws.max_row}")
    
    # Mevcut cari kodlarını kontrol et
    existing_codes = {cari.CariKod for cari in session.query(Cari.CariKod).all()}
    print(f"✓ Veritabanında {len(existing_codes)} cari bulunuyor")
    
    added = 0
    updated = 0
    skipped = 0
    errors = []
    
    # Veriyi oku ve veritabanına ekle (2. satırdan başla)
    for row_idx in range(2, ws.max_row + 1):
        try:
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value
            
            # Zorunlu alanları kontrol et
            cari_kodu = row_data.get('CARİ KODU', '').strip()
            unvan = row_data.get('Unvan', '').strip()
            
            if not cari_kodu or not unvan:
                skipped += 1
                continue
            
            # Mevcut kaydı kontrol et
            existing_cari = session.query(Cari).filter(Cari.CariKod == cari_kodu).first()
            
            # Alan eşlemesi
            vkn = row_data.get('VKN', '').strip() if row_data.get('VKN') else None
            tckn = row_data.get('TCKN', '').strip() if row_data.get('TCKN') else None
            
            if existing_cari:
                # Güncelle
                existing_cari.Unvan = unvan
                existing_cari.VergiDairesi = row_data.get('VergiDairesi', '').strip() if row_data.get('VergiDairesi') else None
                existing_cari.VergiNo = vkn
                existing_cari.Tckn = tckn
                existing_cari.Adres = row_data.get('ADRES', '').strip() if row_data.get('ADRES') else None
                existing_cari.Il = row_data.get('İL', '').strip() if row_data.get('İL') else None
                existing_cari.Ilce = row_data.get('İLÇE', '').strip() if row_data.get('İLÇE') else None
                existing_cari.Ulke = row_data.get('ÜLKE', '').strip() if row_data.get('ÜLKE') else 'Türkiye'
                existing_cari.Eposta = row_data.get('Eposta', '').strip() if row_data.get('Eposta') else None
                existing_cari.Telefon = row_data.get('Tel', '').strip() if row_data.get('Tel') else None
                existing_cari.UpdatedAt = datetime.now()
                updated += 1
            else:
                # Yeni kayıt ekle
                yeni_cari = Cari(
                    CariKod=cari_kodu,
                    Unvan=unvan,
                    CariTip='TUZEL',  # Excel'de şirketi gösterdiği için
                    Rol='MUSTERI',  # Varsayılan olarak müşteri
                    VergiDairesi=row_data.get('VergiDairesi', '').strip() if row_data.get('VergiDairesi') else None,
                    VergiNo=vkn,
                    Tckn=tckn,
                    Adres=row_data.get('ADRES', '').strip() if row_data.get('ADRES') else None,
                    Il=row_data.get('İL', '').strip() if row_data.get('İL') else None,
                    Ilce=row_data.get('İLÇE', '').strip() if row_data.get('İLÇE') else None,
                    Ulke=row_data.get('ÜLKE', '').strip() if row_data.get('ÜLKE') else 'Türkiye',
                    Eposta=row_data.get('Eposta', '').strip() if row_data.get('Eposta') else None,
                    Telefon=row_data.get('Tel', '').strip() if row_data.get('Tel') else None,
                    CreatedAt=datetime.now(),
                    AktifMi=True
                )
                session.add(yeni_cari)
                added += 1
            
            # Her 50 kaydı commit et
            if (added + updated) % 50 == 0:
                session.commit()
                print(f"  ✓ {added + updated} kayıt işlendi...")
                
        except Exception as e:
            errors.append((row_idx, str(e)))
            print(f"  ⚠️ Satır {row_idx} hatasında: {str(e)[:60]}")
    
    # Son commit
    session.commit()
    session.close()
    
    print("\n" + "="*60)
    print("📊 İTHALAT SONUÇLARI:")
    print("="*60)
    print(f"✅ Yeni eklenen: {added}")
    print(f"🔄 Güncellenen: {updated}")
    print(f"⏭️  Atlanılan: {skipped}")
    print(f"❌ Hatalar: {len(errors)}")
    print("="*60)
    
    if errors:
        print("\nHata detayları (ilk 10):")
        for row_idx, error in errors[:10]:
            print(f"  Satır {row_idx}: {error}")

if __name__ == '__main__':
    import_cari_kartlari()
