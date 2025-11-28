"""
DİJİTAL ARŞİV - Excel Rapor Export
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from io import BytesIO


class ArchiveReportExporter:
    """Excel rapor export"""
    
    def export_document_list(self, documents: list, filters: dict = None) -> BytesIO:
        """
        Belge listesini Excel'e export et
        
        Returns:
            BytesIO (Excel dosyası)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Belgeler"
        
        # Header
        headers = [
            "ID", "Belge Tipi", "Kategori", "Durum", "İş Emri No", 
            "Cari", "Yükleyen", "Yüklenme Tarihi", "Onaylayan", "Onay Tarihi",
            "Dosya Adı", "Boyut (MB)"
        ]
        
        ws.append(headers)
        
        # Header styling
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for doc in documents:
            # Work order no
            wo_no = ""
            if hasattr(doc, 'work_order') and doc.work_order:
                wo_no = doc.work_order.wo_number
            
            # Cari title
            cari_title = ""
            if hasattr(doc, 'cari') and doc.cari:
                cari_title = doc.cari.Unvan
            
            # Uploader
            uploader = ""
            if hasattr(doc, 'uploaded_by') and doc.uploaded_by:
                uploader = doc.uploaded_by.full_name
            elif hasattr(doc, 'uploaded_by_portal_user') and doc.uploaded_by_portal_user:
                uploader = doc.uploaded_by_portal_user.full_name
            
            # Approver
            approver = ""
            if hasattr(doc, 'approved_by') and doc.approved_by:
                approver = doc.approved_by.full_name
            
            ws.append([
                doc.id,
                doc.document_type.value,
                doc.category.value,
                doc.status.value,
                wo_no,
                cari_title,
                uploader,
                doc.uploaded_at.strftime("%Y-%m-%d %H:%M"),
                approver,
                doc.approved_at.strftime("%Y-%m-%d %H:%M") if doc.approved_at else "",
                doc.file_name,
                doc.file_size_mb
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_expiry_report(self, expiring_docs: list, expired_docs: list) -> BytesIO:
        """
        Süre sonu raporu Excel'e export
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Dolmak üzere
        ws1 = wb.active
        ws1.title = "Dolmak Üzere"
        
        headers = ["Belge Tipi", "Kategori", "İş Emri No", "Cari", "Geçerlilik Bitiş", "Kalan Gün"]
        ws1.append(headers)
        
        # Header styling
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for item in expiring_docs:
            doc = item['document']
            days = item['days_until_expiry']
            
            ws1.append([
                doc.document_type.value,
                doc.category.value,
                doc.work_order.wo_number if doc.work_order else "",
                doc.cari.Unvan if doc.cari else "",
                doc.expires_at.strftime("%Y-%m-%d"),
                days
            ])
        
        # Sheet 2: Dolmuş
        ws2 = wb.create_sheet("Dolmuş")
        ws2.append(headers)
        
        # Header styling
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for doc in expired_docs:
            ws2.append([
                doc.document_type.value,
                doc.category.value,
                doc.work_order.wo_number if doc.work_order else "",
                doc.cari.Unvan if doc.cari else "",
                doc.expires_at.strftime("%Y-%m-%d"),
                0
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
